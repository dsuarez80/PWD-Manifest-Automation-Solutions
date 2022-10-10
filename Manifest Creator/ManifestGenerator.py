from openpyxl import load_workbook
import os, subprocess
from datetime import datetime, date, timedelta

weekdays = {0 : "MONDAY", 1 : "TUESDAY", 2 : "WEDNESDAY", 3 : "THURSDAY", 4 : "FRIDAY", 5 : "SATURDAY", 6 : "SUNDAY"}
months = {1 : "JANUARY", 2 : "FEBRUARY", 3 : "MARCH", 4 : "APRIL", 5 : "MAY", 6 : "JUNE", 7 : "JULY", 8 : "AUGUST", 9 : "SEPTEMBER", 10 : "OCTOBER", 11 : "NOVEMBER", 12 : "DECEMBER"}
unf_manifests_filepath = "//dc01/_SHARE/Company/Forms/Daily Job sheets/DAILY JOB SHEETS - "

def execute():
    print("How many days ahead of current date?")
    daysn = 0

    while True:
        try:
            choice = int(input())
        except ValueError:
            print("Invalid input.")
            continue

        if choice > 0:
            daysn = choice
            break
        elif choice == 0:
            print("Viewing files for today.")
            break
        else:
            print("No files have been created.")
            exit()

    current_day = date.today()
    future_date = current_day + timedelta(days = daysn)

    weekdayN = future_date.weekday()
    global weekday
    weekday = weekdays[weekdayN]

    formatted_date = date.strftime(future_date, "%m-%d-%Y")
    global active_date
    active_date = formatted_date

    manifests_filepath = get_manifests_filepath()
    if os.path.exists(manifests_filepath):
        print("Manifests filepath:", manifests_filepath)
    else:
        print("Creating Manifests filepath for:\n", manifests_filepath)
        create_manifests_filepath()
    
    print("Creating files for", weekday, active_date)
    global crews
    crews = []

    with open('Manifest Creator/resources/lead_installers.txt') as f:
        lines = f.readlines()
        for line in lines:
            crews.append(line.split())

    for crew in crews:
        generate_template(crew[0])
    
    print("\nWould you like to print these spreadsheets? [Y]es [N]o - Active date:", active_date)

    while True:
        choice = input()[0].lower()

        if choice == 'y':
            print("Printing manifests...")
            print_manifests(active_date)
            break
        elif choice == 'n':
            print("Printing cancelled.")
            break
        else:
            print("Invalid.")

def get_manifests_filepath(d = None):
    if not d:
        d = active_date
    month = months[datetime.strptime(d, "%m-%d-%Y").month]
    year = datetime.strptime(d, "%m-%d-%Y").year

    return unf_manifests_filepath + month + " " + str(year) + "/Install Manifests/"

def create_manifests_filepath(d = None):
    if not d:
        d = active_date
    month = months[datetime.strptime(d, "%m-%d-%Y").month]
    year = datetime.strptime(d, "%m-%d-%Y").year
    newdir = unf_manifests_filepath + month + " " + str(year)
    os.mkdir(newdir)
    newdir = newdir + "/Install Manifests"
    os.mkdir(newdir)

    print("Created new Manifests directory at:\n", newdir)

#creates manifest spreadsheet from passed in manifest object using workbook template
def create_workbook(manifest):
    template_name = "INSTALLATION DAILY MANIFEST NAME DATE"
    workbook_path = "Manifest Creator/resources/" + template_name + ".xlsx"

    if not os.path.exists(workbook_path):
        print("Workbook template does not exist")
        return False
    
    workbook = load_workbook(workbook_path)
    sheet = workbook.active
    
    sheet["A2"] = manifest.date
    sheet["A3"] = "DATE: " + weekdays[datetime.strptime(manifest.date, "%m-%d-%Y").weekday()]
    sheet["D6"] = "LEAD INSTALLER: " + manifest.lead
    sheet["D7"] = "CREW: " + ", ".join(manifest.crew.split())

    for i in range(len(manifest.workorders)):
        cellrow = str(i + 11)
        sheet["A" + cellrow] = manifest.workorders[i].builder
        sheet["B" + cellrow] = manifest.workorders[i].subdivision
        sheet["C" + cellrow] = manifest.workorders[i].lot
        sheet["D" + cellrow] = manifest.workorders[i].windows
        sheet["E" + cellrow] = manifest.workorders[i].doors
        sheet["F" + cellrow] = manifest.workorders[i].notes    
        
    template_name = template_name.split(" ")
    template_name = " ".join(template_name[:-2]) + " " + manifest.lead.upper() + " " + manifest.date + ".xlsx"
    file_path = get_manifests_filepath(manifest.date) + manifest.lead + "/"
    final_path = file_path + template_name

    if not os.path.exists(file_path):
        print("\nDirectory does not exist for:", manifest.lead)
        new_dir = os.path.join(get_manifests_filepath(manifest.date), manifest.lead)
        os.mkdir(new_dir)
        print("Created directory for:", manifest.lead, "under the following filepath:")
        print(new_dir, "\n")
    
    workbook.save(filename = final_path)

def generate_template(lead_name):
    template_name = "INSTALLATION DAILY MANIFEST NAME DATE"
    workbook_path = "Manifest Creator/resources/" + template_name + ".xlsx"
    if not os.path.exists(workbook_path):
        print("Workbook template does not exist.")
        exit()

    workbook = load_workbook(workbook_path)
    sheet = workbook.active

    sheet["A2"] = active_date
    sheet["A3"] = "DATE: " + weekday
    sheet["D6"] = "LEAD INSTALLER: " + lead_name
    teammates = None

    for crew in crews:
        if lead_name == crew[0]:
            if len(crew) > 1:
                teammates = ", ".join(crew[1:])
                sheet["D7"] = "CREW: " + teammates

    template_name = template_name.split(" ")
    template_name = " ".join(template_name[:-2]) + " " + lead_name.upper() + " " + active_date + ".xlsx"
    manifests_filepath = get_manifests_filepath()
    lead_name_path = manifests_filepath + lead_name + "/"
    final_path = lead_name_path + template_name

    if not os.path.exists(lead_name_path):
        print("\nDirectory does not exist for:", lead_name)
        new_dir = os.path.join(manifests_filepath, lead_name)
        os.mkdir(new_dir)
        print("Created directory for:", lead_name, "under the following filepath:")
        print(new_dir, "\n")
    
    if not os.path.exists(final_path):
        workbook.save(filename = final_path)
        print(template_name)
    else:
        print("File already exists at:\n", os.path.normpath(final_path))

        
def print_manifests(givendate):
    manifests_filepath = get_manifests_filepath(givendate)
    for f in os.listdir(manifests_filepath):
        for file in os.listdir(manifests_filepath + f + "/"):
            if(file.endswith(givendate + ".xlsx")):
                workbook = load_workbook(manifests_filepath + f + "/" + file)
                sheet = workbook.active
                A11 = sheet["A11"].value

                if not A11 == None:
                    print("Printing manifest for:", f)
                    os.startfile(os.path.normpath(manifests_filepath + f + "/" + file), "print")
                else:
                    print("Manifest for", f, "is incomplete. Print Aborted.")

def print_manifest(lead, givendate):
    manifests_filepath = get_manifests_filepath(givendate)
    for f in os.listdir(manifests_filepath):
        if f.upper() == lead.upper():
            for file in os.listdir(manifests_filepath + f + "/"):
                if file.endswith(lead + " " + givendate + ".xlsx"):
                    print("Printing manifest file:", file)
                    p = os.path.normpath(manifests_filepath + f + "/" + file)
                    os.startfile(p, "print")
    print()

def open_file(lead, givendate):
    manifests_filepath = get_manifests_filepath(givendate)
    for f in os.listdir(manifests_filepath):
        if f.upper() == lead.upper():
            for file in os.listdir(manifests_filepath + f + "/"):
                if file.endswith(lead + " " + givendate + ".xlsx"):
                    p = os.path.normpath(manifests_filepath + f + "/" + file)
                    subprocess.Popen(r'explorer /select,' + '"' + p + '"')
                    os.startfile(p)
                    
if __name__ == "__main__":
    execute()